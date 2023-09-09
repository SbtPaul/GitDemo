import openpyxl


class HomePageData:
    test_HomePageData = [{"firstname": "subrata", "email": "sbt@gmail.com", "gender": "Male"},
                         {"firstname":"Nobonita", "email": "aa@bb.com", "gender": "Female"}]


    @staticmethod
    def getTestData(testCaseID):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\SudipNobonita\\Desktop\\python\\xlTestData\\testData.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testCaseID:
                for j in range(2, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                # print(Dict)
        return[Dict]