import openpyxl

Dict ={}
book = openpyxl.load_workbook("C:\\Users\\SudipNobonita\\Desktop\\python\\xlTestData\\testData.xlsx")
sheet = book.active
cell = sheet.cell(row=2, column=2)
print(cell.value)
cell2= sheet.cell(row=3, column=2)
print(cell2.value)

#write back on xlsx file
sheet.cell(row=2, column=5).value = "write back on xlsx"
print(sheet.cell(row=2, column=5).value)

print(sheet.max_row)
print(sheet.max_column)

# print all row an column values

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)

#print row for TC=2 only
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "TC2":
        for j in range(2, sheet.max_column+1):
            # print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(Dict)
